import re
from openpyxl import load_workbook

### See README for more information

### put path of file here
nameOfFile = r"file_path_here"

workbook = load_workbook(filename= nameOfFile)
workbook.sheetnames
sheet = workbook.active
sheet
sheet.title


def harmonization():
    """
    Synthesizes adjacent rows that share a common value in column A

    This funciton checks to see if adjacent rows share the same value in
    column A, and if so, combines the values in column B in both rows with a
    "/" into the earliest row and deletes the later row.

    ex:
    A    B                        A    B
    _______       returns ->     ________
    1    a                        1    a
    2    b                        2   b/c
    2    c                        3  d/e/f
    3    d
    3    e
    3    f
    _______                       _______
    """
    i = 1
    while sheet["A" +str(i)].value is not None:
        if i != 1 and i != 0:
            if sheet["A" +str(i)].value == sheet["A" +str(i - 1)].value:
                # print("i is " + str(i))

                cur_letter = "A" +str(i)
                prev_letter = "A" +str(i - 1)

                cur_letter_b= "B" +str(i)
                prev_letter_b = "B" +str(i - 1)

                sheet[prev_letter_b] = str(sheet[prev_letter_b].value)+ "/" + str(sheet[cur_letter_b].value)

                sheet.delete_rows(idx = i)
                i -= 1
        i += 1


def matching_function():
    """
    Matches values in column D to to column B with column C as a reference for
    values in D and A for values for B.

    If values in column C corrospond with respective values in column D, then
    matching_function will match values from D to A into column B given that
    values in column A are codes from column C. If not, then matching_function
    will put a "-" placeholder. This will work for multiple values in column A
    that are seperated with a "/".

    ex:
    A    B    C    D                          A    B    C    D
    __________________       returns ->      __________________
    3         1    ab                         3    ef   1   ab
    2         2    cd                         2    cd   2   cd
    7         3    ef                         7    -    3   ef
   1/2        -     -                       1/2  ab/cd  -    -
    __________________                       __________________
    """
    f = 2
    while sheet["C" +str(f)].value is not None:
        z = 2
        while sheet["A" +str(z)].value is not None:

            x = re.split('/',str(sheet["A" +str(z)].value))
            for q in range(len(x)):
                if x[q] ==  str(sheet["C" + str(f)].value):
                    if sheet["B" + str(z)].value is not None:

                        sheet["B" +str(z)] = str(sheet["B" +str(z)].value)+ "/" + str(sheet["D" + str(f)].value)
                    else:
                        sheet["B" +str(z)] = sheet["D" + str(f)].value
            z +=1
        f += 1
    c = 1
    while sheet["A" +str(c)].value is not None:
        c += 1
    for i in range(c):
        if i > 0:
            if sheet["B" + str(i)].value is None:
                sheet["B" + str(i)] = "-"


def zero_starting_eliminater():
    """
    Finds elements in column A that start with "0" and deletes the "0".

    zero_starting_eliminater parses through column A and eliminates "0" from
    elements that start with 0. This includes elements that have multiple values
    seperated by a "/".

    Note: Excel automatially eleminates numbers starting with 0. To prevent this
    from happening, find the "Numbers Format" option in Excel and select the
    "Text" option for highlighted cells.
    """
    i = 1
    while sheet["A" +str(i)].value is not None:
        # print(i)
        x = re.split('/',str(sheet["A" +str(i)].value))
        if len(x) > 1:
            final = ""
            res = []
            [res.append(q) for q in x if q not in res]
            for b in range(len(res)):
                if b == 0:
                    if str(res[b][0]) == "0":
                        final = final  + res[b][1:]
                    else:
                        final = final + res[b]
                else:
                    if str(res[b][0]) == "0":
                        final = final +"/" + res[b][1:]
                    else:
                        final = final +"/" + res[b]
            sheet["A" +str(i)] = final
        else:
            if str(x[0][0]) == "0":
                sheet["A" + str(i)] = x[0][1:]
        i += 1


### call desired functions here:


# harmonization()
# matching_function()
# zero_starting_eliminater()


workbook.save(filename= nameOfFile)
